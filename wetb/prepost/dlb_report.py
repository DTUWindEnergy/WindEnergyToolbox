# -*- coding: utf-8 -*-
"""
Created on Fri Aug 30 14:03:27 2024

@author: nicgo
"""

import numpy as np
import os
import glob
import openpyxl
from openpyxl.styles import Alignment
from wetb.hawc2.Hawc2io import ReadHawc2
from wetb.gtsdf import load_statistic
from wetb.utils.envelope import projected_extremes
from wetb.utils.rotation import projection_2d
from wetb.fatigue_tools.fatigue import eq_load

def get_filenames(path, ext):
    roots = [path] + [os.path.join(path, f) for f in os.listdir(path) if os.path.isdir(os.path.join(path, f))]
    return sorted([filename for root in roots for filename in glob.glob(os.path.join(root, ext))])

def get_statistic(time, data, statistics=['min', 'mean', 'max', 'std', 'eq3', 'eq4', 'eq6', 'eq8', 'eq10', 'eq12'], no_bins=46):
    def get_stat(stat):
        if hasattr(np, stat):
            return getattr(np, stat)(data, 0)
        elif (stat.startswith("eq") and stat[2:].isdigit()):
            from wetb.fatigue_tools.fatigue import eq_load
            m = float(stat[2:])
            return [eq_load(sensor, no_bins, m, time[-1] - time[0] + time[1] - time[0])[0][0] for sensor in data.T]
    return np.array([get_stat(stat) for stat in statistics]).T

def dlb_postprocessing(path, ext,
                       sensor_lst=["Ae rot. torque", "Ae rot. power", "Ae rot. thrust", "bea1 angle_speed"],
                       statistics=['min', 'mean', 'max', 'std', 'eq3', 'eq4', 'eq6', 'eq8', 'eq10', 'eq12'],
                       load_sensor_list=[], angles=np.linspace(-150, 180, 12), sweep_angle=30, degrees=True,
                       no_bins=46, m=4, neq=int(1e7)):
    filenames = get_filenames(path, ext)
    dlb_statistics = {}
    dlb_extreme_loads = {}
    dlb_fatigue_loads = {}
    if ext == "*.hdf5":
        for f in filenames:
            file_statistics = load_statistic(f, xarray=True)
            dlb_statistics[os.path.basename(os.path.splitext(f)[0])] = {}
            for s in sensor_lst:
                dlb_statistics[os.path.basename(os.path.splitext(f)[0])][s] = {}
                for stat in file_statistics['stat'].values:
                    dlb_statistics[os.path.basename(os.path.splitext(f)[0])][s][stat] = file_statistics.sel(sensor_name = s, stat = stat).values[()]
            res_file = ReadHawc2(f)
            data = res_file.ReadAll()
            extreme_loads = {}
            fatigue_loads = {}
            for s, ix, iy in load_sensor_list:
                extremes = projected_extremes(np.vstack([data[:,ix], data[:,iy]]).T, angles, sweep_angle, degrees)
                extreme_loads[s] = {}
                fatigue_loads[s] = {}
                for a in range(len(angles)):
                    extreme_loads[s][angles[a]] = extremes[a][1]
                    fatigue_loads[s][angles[a]] = eq_load(data[:,[ix,iy]] @ projection_2d(angles[a], degrees), no_bins=no_bins, m=m, neq=neq)[0][0]
            dlb_extreme_loads[os.path.basename(os.path.splitext(f)[0])] = extreme_loads 
            dlb_fatigue_loads[os.path.basename(os.path.splitext(f)[0])] = fatigue_loads               
    elif ext == "*.sel":
        for f in filenames:
            res_file = ReadHawc2(f)
            time = res_file.t
            data = res_file.ReadAll()
            info = res_file.ChInfo[0]
            file_statistics = get_statistic(time, data, statistics)
            dlb_statistics[os.path.basename(os.path.splitext(f)[0])] = {}
            for s in sensor_lst:
                dlb_statistics[os.path.basename(os.path.splitext(f)[0])][s] = {}
                for st in statistics:
                    dlb_statistics[os.path.basename(os.path.splitext(f)[0])][s][st] = file_statistics[info.index(s), statistics.index(st)]
            extreme_loads = {}
            fatigue_loads = {}
            for s, ix, iy in load_sensor_list:
                extremes = projected_extremes(np.vstack([data[:,ix], data[:,iy]]).T, angles, sweep_angle, degrees)
                extreme_loads[s] = {}
                fatigue_loads[s] = {}
                for a in range(len(angles)):
                    extreme_loads[s][angles[a]] = extremes[a][1]
                    fatigue_loads[s][angles[a]] = eq_load(data[:,[ix,iy]] @ projection_2d(angles[a], degrees), no_bins=no_bins, m=m, neq=neq)[0][0]
            dlb_extreme_loads[os.path.basename(os.path.splitext(f)[0])] = extreme_loads 
            dlb_fatigue_loads[os.path.basename(os.path.splitext(f)[0])] = fatigue_loads 
    return dlb_statistics, dlb_extreme_loads, dlb_fatigue_loads
    
def write_dlb_report_to_excel(dlb_statistics, dlb_extreme_loads, dlb_fatigue_loads, excel_filename):
    workbook = openpyxl.Workbook()
    contents = {"Statistics": dlb_statistics, "Extreme loads": dlb_extreme_loads, "Fatigue loads": dlb_fatigue_loads}
    for sheet_name, sheet_values in contents.items():
        sheet = workbook.create_sheet(sheet_name)
        central_headers = ['Simulation']
        inner_headers = ['']
        first_item = next(iter(sheet_values.values()))
        if isinstance(first_item, dict):
            for central_key, inner_dict in first_item.items():
                if isinstance(inner_dict, dict):
                    for inner_key in inner_dict.keys():
                        central_headers.append(central_key)
                        inner_headers.append(inner_key)
                else:
                    central_headers.append(central_key)
                    inner_headers.append('')
        for col_num, header in enumerate(central_headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)
        for col_num, header in enumerate(inner_headers, start=1):
            sheet.cell(row=2, column=col_num, value=header)
        current_central = central_headers[0]
        start_col = 1
        for col_num, central_header in enumerate(central_headers[1:], start=2):
            if central_header != current_central:
                if col_num - start_col > 1:
                    sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_num-1)
                start_col = col_num
                current_central = central_header
        if len(central_headers) - start_col > 0:
            sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=len(central_headers))
        for col_num in range(1, len(central_headers) + 1):
            sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
        for row_num, (outer_key, central_dict) in enumerate(sheet_values.items(), start=3):
            sheet.cell(row=row_num, column=1, value=outer_key)
            col_num = 2
            for central_key, inner_dict in central_dict.items():
                if isinstance(inner_dict, dict):
                    for inner_key, inner_value in inner_dict.items():
                        sheet.cell(row=row_num, column=col_num, value=inner_value)
                        col_num += 1
                else:
                    sheet.cell(row=row_num, column=col_num, value=inner_dict)
                    col_num += 1
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    workbook.save(excel_filename)

if __name__ == '__main__':        
    folder_path = r"C:\Users\nicgo\Documents\Turbine_Models\WindEnergyToolbox\res_hawc_ascii"
    dlb_statistics, dlb_extreme_loads, dlb_fatigue_loads = dlb_postprocessing(folder_path, "*.sel", load_sensor_list=[('Tower base bending moment', 16, 17), ('Blade 1 root bending moment', 25, 26), ('Blade 2 root bending moment', 28, 29), ('Blade 3 root bending moment', 31, 32)])
    excel_filename = 'DLB_Statistics_Report_ascii.xlsx'
    write_dlb_report_to_excel(dlb_statistics, dlb_extreme_loads, dlb_fatigue_loads, excel_filename)
