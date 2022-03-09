# -*- coding: utf-8 -*-
"""
Created on Fri Nov 20 10:11:06 2015

@author: tlbl
"""
# arctan and pi are required because they are in the formulas that are
# evaluated
from numpy import (floor, arctan, pi, log, log10, sin, cos, tan, e, arcsin,
                   arccos)
import pandas as pd
from openpyxl import load_workbook
from argparse import ArgumentParser
import os


def multi_for(iterables):
    """
    Routine to create list with combination of elements.
    """
    if not iterables:
        yield ()
    else:
        for item in iterables[0]:
            for rest_tuple in multi_for(iterables[1:]):
                yield (item,) + rest_tuple


class GeneralDLC(object):
    """
    Basic class to generate the DLC spreadsheets. It contains routines to
    handle the different types of tags.

    * Constants: are fixed in the current DLC, e.g. reference turbulence\
        intensity, rotor radius, reference wind speed, ....

    * Variables: define the number of cases in a DLC through their combination\
        e.g. wind speed, number of turbulence seeds, yaw angle, ....


    * Functions: depend on other tags e.g turbulence intensity, file name, ....

    """

    def __init__(self):

        pass

    def remove_from_dict(self, non_defaults, defaults):

        for key in non_defaults.keys():
            try:
                del defaults[key]
            except:
                pass
        return defaults

    def add_variables_tag(self, dlc, variables, variables_order):

        cases_len = []
        for tag in variables_order:
            dlc[tag] = []
            v = variables[tag]
            for i in range(len(v)-1):
                try:
                    v.remove('')
                except:
                    pass
            if tag == '[seed]':
                cases_len.append(int(v[0]))
            elif tag == '[wave_seed]':
                cases_len.append(int(v[0]))
            else:
                cases_len.append(len(v))
        cases_index = multi_for(list(map(range, cases_len)))

        # when no seeds are used, otherwise i_seed is not set
        i_seed = -1
        if '[wsp]' in variables_order:
            i_wsp = variables_order.index('[wsp]')
            len_wsp = len(variables['[wsp]'])
        else:
            raise ValueError('Missing VARIABLE (V) [wsp] tag!')
        if '[seed]' in variables_order:
            i_seed = variables_order.index('[seed]')
        if '[wave_seed]' in variables_order:
            i_wave_seed = variables_order.index('[wave_seed]')
        if i_seed > i_wsp:
            raise ValueError('column [seed] should come BEFORE [wsp] !!')

        for irow, row in enumerate(cases_index):
            for icol, col in enumerate(row):
                if variables_order[icol] == '[seed]':
                    counter = floor(irow/len_wsp) + 1
                    value = '%4.4i' % (1000*counter + row[i_wsp] + 1)
                elif variables_order[icol] == '[wave_seed]':
                    # FIXME: shouldn't we also have unique wave seeds??
                    # that is not the case with this implementation
                    value = '%4.4i' % (100*(row[i_wsp]+1) + row[i_wave_seed] + 1)
#                    value = '%4.4i' % (1000*counter + row[i_wsp] + 101)
#                    value = '%4.4i' % (irow+1)
#                    value = '%4.4i' % (10000*(row[i_wave_dir])] + 1) + \
#                                        1000*(row[i_Hs])] + 1) + \
#                                        10*(row[i_Tp])] + 1) +\
#                                        row[i_seed])] + 1)

                else:
                    value = variables[variables_order[icol]][col]
                    if not isinstance(value, float) and not isinstance(value, int):
                        value = str(value)
                dlc[variables_order[icol]].append(value)

    def add_constants_tag(self, dlc, constants):

        for key in constants.keys():
            dlc[key] = [constants[key]]*len(dlc['[wsp]'])

    def sort_formulas(self, formulas):
        # sort formulas based on their dependency
        keys_list = sorted(formulas)
        for i in range(len(keys_list)):
            for ikey, key in enumerate(keys_list):
                formula = formulas[key]
                for ikey2, key2 in enumerate(keys_list):
                    if key2 in formula:
                        if ikey < ikey2:
                            keys_list.pop(ikey)
                            keys_list.insert(ikey2, key)
                            break
        return keys_list

    def eval_formulas(self, dlc):

        for key in dlc.keys():
            if isinstance(dlc[key][0], str):
                if "[" in dlc[key][0]:
                    for key2 in dlc.keys():
                        for iformula, formula in enumerate(dlc[key]):
                            if key2 in formula:
                                dlc[key][iformula] = dlc[key][iformula].replace(key2, '%s'%dlc[key2][iformula])
                    for iformula, formula in enumerate(dlc[key]):
                        formula = formula.replace(',', '.')
                        formula = formula.replace(';', ',')
                        dlc[key][iformula] = eval(formula)

    def add_formulas(self, dlc, formulas):

        keys_list = self.sort_formulas(formulas)

        # specify the precision of the tag as used in the formulas
        # this does NOT affect the precision of the tag itself, only when used
        # in a formula based tag.
        formats = {'[wsp]':'%02i', '[gridgustdelay]':'%02i',
                   '[wdir]':'%03i', '[G_phi0]':'%03i',
                   '[sign]':'%s',
                   '[Hs]':'%05.02f', '[Tp]':'%05.02f'}

        for fkey in keys_list:
            flist = []
            for i in range(len(dlc['[wsp]'])):
                formula = formulas[fkey]
                for key in dlc.keys():
                    if key in formula:
                        if formula[0] == '"':
                            try:
                                fmt = formats[key]
                            except KeyError:
                                fmt = '%04i'
                            try:
                                value = float(dlc[key][i])
                            except ValueError:
                                # this is for string tags
                                value = dlc[key][i]
                                fmt = '%s'
                            formula = formula.replace(key, fmt % value)
                        elif key in formula:
                            formula = formula.replace(key, '%s' % dlc[key][i])
                formula = formula.replace(',', '.')
                formula = formula.replace(';', ',')
                formula = formula.replace('\n', ' ')
                try:
                    flist.append(eval(formula))
                except Exception as exc:
                    print('following formula failed to execute:')
                    print(formula)
                    print('and raised the following exception:')
                    raise exc

            dlc[fkey] = flist


class GenerateDLCCases(GeneralDLC):
    """
    Class to generate Excell sheets for each DLB case starting from a single
    Excell sheet.

    Parameters
    ----------

    filename: str
        Name of the excel spreadsheet containing the definition of all the
        cases to generate.

    folder: str
        Name of the folder in which to save the DLB cases.

    Example
    -------
        DLB = GenerateDLCCases()
        DLB.execute()


    """

    def execute(self, filename='DLCs.xlsx', folder='', isheets=None):

        book = load_workbook(filename, data_only=True)

        # Read all the initialization constants and functions of the main sheet
        # The main sheet is assumed to be the first one (i=0)
        general_constants = {}
        general_functions = {}
        sheet = book.worksheets[0]
        # note that sheet.cell(i,j) is using 1-based indexing (row nr 1)
        # first column is just for readability
        for colnr in range(2, sheet.max_column+1):
            if sheet.cell(10, colnr).value != None:
                general_constants[str(sheet.cell(10, colnr).value)] = \
                    sheet.cell(11, colnr).value
            if sheet.cell(14, colnr).value != None:
                general_functions[str(sheet.cell(14, colnr).value)] = \
                    sheet.cell(15, colnr).value

        if isheets is None:
            # refer to sheet number, so 1-based indexing
            isheets = list(range(1,len(book.sheetnames)))

        # Loop through all the sheets. Each sheet correspond to a DLC.
        for isheet in isheets:

            sheet = book.worksheets[isheet]

            print('Sheet #%i' % isheet, book.sheetnames[isheet])
            # Read the actual sheet.
            constants = {}
            variables = {}
            formulas = {}
            variables_order = []
            # Loop through the columns
            for i in range(1,sheet.max_column+1):
                if sheet.cell(1, i).value is not None:
                    tag = str(sheet.cell(1+1, i).value)
                    if len(tag) > 0:
                        # FIXME: only works if [wsp] is defined as variable
                        # and [seed] tags are present
                        if sheet.cell(1, i).value == 'C':
                            constants[tag] = sheet.cell(3, i).value
                        if sheet.cell(1, i).value == 'V':
                            variables_order.append(tag)
                            variables[tag] = \
                                [sheet.cell(j, i).value for j in range(3, sheet.max_row+1) if sheet.cell(j, i).value != None]
                        if sheet.cell(1, i).value == 'F':
                            formulas[tag] = str(sheet.cell(3, i).value)
            dlc = {}

            # make copies of the general constants and functions since remove
            # will otherwise remove them for the following sheets as well
            sheet_gen_con = self.remove_from_dict(variables, general_constants.copy())
            sheet_gen_con = self.remove_from_dict(constants, sheet_gen_con)
            sheet_gen_fun = self.remove_from_dict(formulas, general_functions.copy())

            self.add_variables_tag(dlc, variables, variables_order)
            self.add_constants_tag(dlc, sheet_gen_con)
            self.add_constants_tag(dlc, constants)
            self.add_formulas(dlc, formulas)
            self.add_formulas(dlc, sheet_gen_fun)
            # TODO: before eval, check if all tags in formula's are present
            self.eval_formulas(dlc)
            df = pd.DataFrame(dlc)
            os.makedirs(folder, exist_ok=True)
            df.to_excel(os.path.join(folder, book.sheetnames[isheet]+'.xlsx'), index=False)


if __name__ == '__main__':

    parser = ArgumentParser(description = "generator of DLB spreadsheets")
    parser.add_argument('--master', type=str, default='DLCs.xlsx', action='store',
                        dest='filename', help='Master spreadsheet file')
    parser.add_argument('--folder', type=str, default='', action='store',
                        dest='folder', help='Destination folder name')
    opt = parser.parse_args()
    DLB = GenerateDLCCases()
    DLB.execute(filename=opt.filename, folder=opt.folder)
