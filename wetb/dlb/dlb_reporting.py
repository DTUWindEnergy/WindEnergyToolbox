import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import os
from wetb.dlb.dlb_postprocs import get_DLC

def dataarray_to_excel(dataarray, path):
    """
    Generate excel file from a DataArray.

    Parameters
    ----------
    dataarray : xarray.DataArray
        DataArray containing data from the DLB.
    path : str
        Path to the new excel file to be created

    Returns
    -------
    None.

    """
    df = dataarray.to_dataframe('value').reset_index()    
    df.to_excel(path, index=False)

def DLB_extreme_loads_to_excel(DLB_extreme_loads, path):
    """
    Generate excel report of DLB extreme loads.

    Parameters
    ----------
    DLB_extreme_loads : xarray.DataArray
        DataArray containing the DLB extreme loading matrices for each load sensor.
    path : str
        Path to the new excel file to be created

    Returns
    -------
    None.

    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sensor in DLB_extreme_loads.coords['sensor_name'].values:
            df = DLB_extreme_loads.sel(sensor_name=sensor).to_pandas()
            df['group'] = DLB_extreme_loads.sel(sensor_name=sensor)['group'].values
            df.to_excel(writer, sheet_name=sensor, index_label='driver')
    workbook = load_workbook(path)
    highlight = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for sensor in DLB_extreme_loads.coords['sensor_name'].values:
        for i in range(12):
            workbook[sensor].cell(row=i + 2, column=int(i/2) + 2).fill = highlight
        workbook[sensor].cell(row=14, column=8).fill = highlight
        workbook[sensor].cell(row=15, column=10).fill = highlight
    workbook.save(path)

def plot_DLB_directional_extreme_loads(DLB_extreme_loads, folder, extension='.png'):
    """
    Plot the DLB directional extreme loads and save them to a given folder.

    Parameters
    ----------
    DLB_extreme_loads : xarray.DataArray ('sensor_name', 'angle')
        DataArray containing the extreme loads of the DLB.
    folder : str
        Path to the folder where the plots will be saved
    extension: str, optional
        Extension of the plot files. The default is '.png'.

    Returns
    -------
    None.

    """
    angles = DLB_extreme_loads.coords['angle'].values
    for i in range(DLB_extreme_loads.shape[0]):
        x = [float(DLB_extreme_loads[i, j])*np.cos(np.deg2rad(angles[j]))
             for j in range(len(angles))]
        y = [float(DLB_extreme_loads[i, j])*np.sin(np.deg2rad(angles[j]))
             for j in range(len(angles))]
        DLC = [get_DLC(DLB_extreme_loads[i, j].group.values[()]) for j in range(len(angles))]
        plt.scatter(x, y)
        [plt.plot([0, x[k]], [0, y[k]], color='black') for k in range(len(x))]
        plt.xlabel('Mx')
        plt.ylabel('My')
        plt.title(DLB_extreme_loads[i].coords['sensor_name'].values[()])
        plt.axhline(0, color='black',linewidth=1)  
        plt.axvline(0, color='black',linewidth=1)  
        plt.xlim(-max(abs(min(x)), abs(max(x)))*1.2, max(abs(min(x)), abs(max(x)))*1.2)
        plt.ylim(-max(abs(min(y)), abs(max(y)))*1.2, max(abs(min(y)), abs(max(y)))*1.2)
        for j in range(len(x)):
            plt.annotate(DLC[j], (x[j], y[j]), textcoords="offset points", xytext=(0,10), ha='center')
        plt.savefig(os.path.join(folder, 'Extreme_' + DLB_extreme_loads[i].coords['sensor_name'].values[()] + extension))
        plt.show()
        
def plot_DLB_directional_equivalent_loads(DLB_fatigue_loads, folder, extension='.png'):
    """
    Plot the DLB directional equivalent loads and save them to a given folder.

    Parameters
    ----------
    DLB_fatigue_loads : xarray.DataArray ('sensor_name', 'angle', 'm')
        DataArray containing the fatigue loads of the DLB. It matches the
        output from get_DLB_fatigue_loads
    folder : str
        Path to the folder where the plots will be saved
    extension: str, optional
        Extension of the plot files. The default is '.png'.

    Returns
    -------
    None.

    """
    m_list = DLB_fatigue_loads.coords['m'].values
    angles = DLB_fatigue_loads.coords['angle'].values
    for i in range(DLB_fatigue_loads.shape[0]):
        for j in range(len(m_list)):
            x = [float(DLB_fatigue_loads[i, k, j])*np.cos(np.deg2rad(angles[k]))
                 for k in range(len(angles))]
            y = [float(DLB_fatigue_loads[i, k, j])*np.sin(np.deg2rad(angles[k]))
                 for k in range(len(angles))]
            plt.scatter(x, y, label='m = ' + str(m_list[j]))
            [plt.plot([0, x[k]], [0, y[k]], color='black') for k in range(len(angles))]
        plt.xlabel('Mx')
        plt.ylabel('My')
        plt.title(DLB_fatigue_loads[i].coords['sensor_name'].values[()])
        plt.axhline(0, color='black',linewidth=1)
        plt.axvline(0, color='black',linewidth=1)
        plt.xlim(-max(abs(min(x)), abs(max(x)))*1.2, max(abs(min(x)), abs(max(x)))*1.2)
        plt.ylim(-max(abs(min(y)), abs(max(y)))*1.2, max(abs(min(y)), abs(max(y)))*1.2)
        plt.legend()
        plt.savefig(os.path.join(folder, 'Fatigue_' + DLB_fatigue_loads[i].coords['sensor_name'].values[()] + extension))
        plt.show()